import streamlit as st
import re
import io
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

st.set_page_config(page_title="AUB Statement to Excel", layout="wide")
st.title("AUB Bank Statement to Excel Converter")
st.write(
    "Upload an Atlantic Union Bank **statement PDF** to convert it into a "
    "structured Excel file with deposits, withdrawals, and running balances."
)

# ── Apple-inspired color palette ──────────────────────────────────────────────
APPLE_BLUE = "0071E3"
APPLE_DEEP_BLUE = "0055CC"
APPLE_BRIGHT_BLUE = "1A8CFF"
APPLE_NEAR_BLACK = "1D1D1F"
APPLE_LIGHT_GRAY = "F5F5F7"
APPLE_MID_GRAY = "D2D2D7"
APPLE_SEC_GRAY = "6E6E73"
WHITE = "FFFFFF"

# ── Helpers ───────────────────────────────────────────────────────────────────

def clean_noise(text):
    """Remove barcode/OCR noise lines from extracted text."""
    MONTHS_PAT = r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)'
    lines = text.split("\n")
    cleaned = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        # If a line has barcode noise BEFORE a transaction, extract just the transaction
        # e.g. "AHPEHKGK... Jan 12 UNITEDHEALTHCARE/BILLING 460.26 92,517.59"
        if not re.match(rf'^{MONTHS_PAT}\s+\d{{1,2}}\s+', stripped):
            embedded = re.search(rf'({MONTHS_PAT}\s+\d{{1,2}}\s+.+)', stripped)
            if embedded and re.search(r'[\d,]+\.\d{2}', embedded.group(1)):
                stripped = embedded.group(1).strip()

        # Skip pure barcode-like strings (long alphanumeric with no spaces)
        if re.match(r'^[A-Z]{10,}$', stripped):
            continue
        if re.match(r'^[A-Z ]{30,}$', stripped) and not any(c.isdigit() for c in stripped):
            if not any(kw in stripped for kw in [
                'TRANSACTION', 'CHECK', 'BALANCE', 'DEPOSIT', 'FUND', 'TRANSFER',
                'CORNERSTONES', 'FAIRFAX', 'PAYCOM', 'FLORES', 'SCHWAB', 'FIDELITY',
                'RELATIONSHIP', 'FLEX BUSINESS', 'ACCOUNT', 'ENDING', 'BEGINNING',
                'TRNSFR', 'MERCHAN', 'FUNDRAISE', 'BIDKIT', 'MONTHLY', 'INTEREST',
                'FEE RECAP', 'SUMMARY', 'SERVICE', 'AVERAGE', 'MINIMUM',
                'COMMONWEALTH', 'SENTARA', 'RAISERIGHT', 'STARKIST', 'AMERICA',
                'DISTRICT', 'UNITEDHEALTHCARE', 'JOHN HANCOCK', 'ROBERT', 'PAYPAL',
                'NETWORK', 'BBGF', 'BTI360', 'NEW YORK', 'WEX INC', 'METKC',
                'UWNCA', 'JK GROUP', 'M&T COMM', 'PPAY'
            ]):
                continue
        # Skip page reference lines
        if re.match(r'^\d{8}\s+\d{7}\s+\d{4}-\d{4}', stripped):
            continue
        if re.match(r'^\d{8}$', stripped):
            continue
        if re.match(r'^\d{7}$', stripped):
            continue
        if re.match(r'^\d{4}-\d{4}$', stripped):
            continue
        if re.match(r'^\d{8}\s+M\d+DDA', stripped):
            continue
        # Clean inline barcode noise (space-separated single chars like "A A L B H N...")
        stripped = re.sub(r'(?:\b[A-Z]\s){5,}[A-Z]\b', '', stripped).strip()
        if not stripped:
            continue
        cleaned.append(stripped)
    return "\n".join(cleaned)


def parse_amount(text):
    """Parse a dollar amount string, returning float. Handles $1,234.56 format."""
    if not text:
        return None
    text = text.strip().replace("$", "").replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def extract_balance_summary(full_text):
    """Extract key balance summary fields from the statement."""
    info = {}

    m = re.search(r'Account Number[:\s]+(\d+)', full_text)
    if m:
        info['account_number'] = m.group(1)

    m = re.search(r'Account Owner\(s\):\s*(.+)', full_text)
    if m:
        info['account_name'] = m.group(1).strip()
    elif re.search(r'CORNERSTONES,?\s*INC', full_text):
        info['account_name'] = 'CORNERSTONES, INC.'

    m = re.search(r'Statement Date\s+(\d{2}/\d{2}/\d{4})', full_text)
    if m:
        info['statement_date'] = m.group(1)

    m = re.search(r'Statement Thru Date\s+(\d{2}/\d{2}/\d{4})', full_text)
    if m:
        info['statement_thru'] = m.group(1)

    m = re.search(r'Account Type.*?Account Number.*?\n(.+?)\s+\d{10}', full_text)
    if m:
        info['account_type'] = m.group(1).strip()
    else:
        m2 = re.search(r'FLEX BUSINESS CKING PLUS', full_text)
        if m2:
            info['account_type'] = 'FLEX BUSINESS CKING PLUS'

    m = re.search(r'Beginning Balance as of\s*(\d{2}/\d{2}/\d{4})\s+\$([\d,]+\.\d{2})', full_text)
    if m:
        info['begin_date'] = m.group(1)
        info['begin_balance'] = parse_amount(m.group(2))

    m = re.search(r'Ending Balance as of\s*(\d{2}/\d{2}/\d{4})\s+\$([\d,]+\.\d{2})', full_text)
    if m:
        info['end_date'] = m.group(1)
        info['end_balance'] = parse_amount(m.group(2))

    m = re.search(r'\+\s*Deposits and Credits\s+\((\d+)\)\s+\$([\d,]+\.\d{2})', full_text)
    if m:
        info['deposit_count'] = int(m.group(1))
        info['deposit_total'] = parse_amount(m.group(2))

    m = re.search(r'-\s*Withdrawals and Debits\s+\((\d+)\)\s+\$([\d,]+\.\d{2})', full_text)
    if m:
        info['withdrawal_count'] = int(m.group(1))
        info['withdrawal_total'] = parse_amount(m.group(2))

    return info


def parse_transactions(full_text):
    """Parse all transactions from the TRANSACTION DETAIL sections."""
    transactions = []

    # Find all TRANSACTION DETAIL sections (skip CHECK TRANSACTION SUMMARY)
    sections = re.split(r'CHECK TRANSACTION SUMMARY', full_text)[0]

    # Match transaction lines: Date Description [Deposit] [Withdrawal] Balance
    # Date format: "Jan DD" or "Feb DD"
    # Amounts: comma-separated with decimals
    lines = sections.split("\n")

    MONTH_MAP = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }

    # Pattern for a transaction line starting with a date
    date_pat = re.compile(
        r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{1,2})\s+'
    )

    # Pattern for amounts at the end of a line
    # Could be: deposit balance | withdrawal balance | just balance
    amount_pat = re.compile(
        r'([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$'
    )
    single_amount_pat = re.compile(
        r'([\d,]+\.\d{2})\s*$'
    )

    i = 0
    year = None

    # Determine year from statement date
    m = re.search(r'Statement Date\s+\d{2}/\d{2}/(\d{4})', full_text)
    if m:
        year = int(m.group(1))
    else:
        year = 2026

    while i < len(lines):
        line = lines[i].strip()

        # Skip headers, footers, and non-transaction lines
        if not line or line.startswith('Account Number') or line.startswith('Statement'):
            i += 1
            continue
        if line.startswith('Page ') or line.startswith('TRANSACTION DETAIL'):
            i += 1
            continue
        if line.startswith('Date') and 'Description' in line:
            i += 1
            continue
        if 'BEGINNING BALANCE' in line:
            i += 1
            continue
        if 'ENDING BALANCE' in line:
            i += 1
            continue
        if line.startswith('RELATIONSHIP') or line.startswith('Balance Summary'):
            i += 1
            continue
        if line.startswith('Beginning Balance as of') or line.startswith('+ Deposits'):
            i += 1
            continue
        if line.startswith('- Withdrawals') or line.startswith('Ending Balance'):
            i += 1
            continue
        if line.startswith('Service Charges') or line.startswith('Average'):
            i += 1
            continue
        if line.startswith('Minimum Balance') or line.startswith('Interest'):
            i += 1
            continue
        if line.startswith('Annual Percentage') or line.startswith('Number of Days'):
            i += 1
            continue
        if line.startswith('Earnings Summary') or line.startswith('Interest Paid'):
            i += 1
            continue
        if line.startswith('Customer Service') or line.startswith('Check/Items'):
            i += 1
            continue
        if any(line.startswith(x) for x in [
            'PO Box', 'Customer Care', 'Monday', 'Saturday', 'Mailing',
            'Glen Allen', 'Visit Us', 'Atlantic', 'Follow us'
        ]):
            i += 1
            continue

        date_match = date_pat.match(line)
        if not date_match:
            i += 1
            continue

        month_str = date_match.group(1)
        day = int(date_match.group(2))
        month = MONTH_MAP.get(month_str, 1)

        # Determine year for this transaction
        txn_year = year
        # If statement is for January but month is Dec, it's previous year
        if m_stmt := re.search(r'Statement Date\s+(\d{2})', full_text):
            stmt_month = int(m_stmt.group(1))
            if month > stmt_month + 1:
                txn_year = year - 1

        rest = line[date_match.end():]

        # Try to find two amounts at end (deposit/withdrawal + balance)
        two_amt = amount_pat.search(rest)
        one_amt = single_amount_pat.search(rest)

        description = ""
        deposit = None
        withdrawal = None
        balance = None

        if two_amt:
            desc_part = rest[:two_amt.start()].strip()
            amt1 = parse_amount(two_amt.group(1))
            amt2 = parse_amount(two_amt.group(2))

            description = desc_part
            balance = amt2
            # Store amount as deposit initially; fix_deposit_withdrawal_classification
            # will use the running balance chain to correctly classify each transaction
            deposit = amt1

        elif one_amt:
            # Only balance (like BEGINNING BALANCE line) - skip these
            desc_part = rest[:one_amt.start()].strip()
            # This might be a line with only a balance and no amount
            # Or it could be that deposit/withdrawal is the single amount
            # Check if it's really just a balance line
            if 'BALANCE' in desc_part.upper():
                i += 1
                continue
            # Single amount could mean it's a deposit with no separate balance shown
            # This is unusual in this format - skip
            i += 1
            continue
        else:
            i += 1
            continue

        # Collect continuation description lines
        j = i + 1
        while j < len(lines):
            next_line = lines[j].strip()
            if not next_line:
                j += 1
                continue
            # If next line starts with a date, stop
            if date_pat.match(next_line):
                break
            # If next line is a header/footer, stop
            if next_line.startswith('Account Number') or next_line.startswith('Page '):
                break
            if next_line.startswith('TRANSACTION DETAIL'):
                break
            if next_line.startswith('Date') and 'Description' in next_line:
                break
            if next_line.startswith('CHECK TRANSACTION'):
                break
            if re.match(r'^\d{8}\s', next_line):
                j += 1
                continue
            # Check if it's a continuation line (no amounts, just text)
            if not re.search(r'[\d,]+\.\d{2}', next_line):
                # It's a description continuation
                description += " " + next_line
                j += 1
            else:
                break
        i = j

        date_str = f"{month:02d}/{day:02d}/{txn_year}"

        transactions.append({
            'date': date_str,
            'description': description.strip(),
            'deposit': deposit,
            'withdrawal': withdrawal,
            'balance': balance,
        })

    return transactions


def fix_deposit_withdrawal_classification(transactions, begin_balance):
    """
    Use the running balance to correctly classify deposits vs withdrawals.
    If balance increases from previous, the amount is a deposit.
    If balance decreases, it's a withdrawal.
    """
    prev_balance = begin_balance
    for txn in transactions:
        amt = txn['deposit'] if txn['deposit'] is not None else txn['withdrawal']
        if amt is None:
            prev_balance = txn['balance'] if txn['balance'] is not None else prev_balance
            continue

        if txn['balance'] is not None and prev_balance is not None:
            diff = round(txn['balance'] - prev_balance, 2)
            if abs(diff - amt) < 0.02:
                # It's a deposit (balance went up by amt)
                txn['deposit'] = amt
                txn['withdrawal'] = None
            elif abs(diff + amt) < 0.02:
                # It's a withdrawal (balance went down by amt)
                txn['withdrawal'] = amt
                txn['deposit'] = None
            elif diff > 0:
                txn['deposit'] = amt
                txn['withdrawal'] = None
            else:
                txn['withdrawal'] = amt
                txn['deposit'] = None

        prev_balance = txn['balance'] if txn['balance'] is not None else prev_balance

    return transactions


def categorize_withdrawal(desc):
    """Categorize a withdrawal by type for the purch1 sheet."""
    d = desc.upper()
    if d.startswith('CHECK #') or d.startswith('CHECK#'):
        return 'CHECKS'
    if 'TRNSFR TO ACCOUNT' in d:
        return 'SWEEPS TO 7459'
    return 'OTHER DEBITS'


def categorize_deposit(desc):
    """Categorize a deposit by type for the dep1 sheet."""
    d = desc.upper()
    if 'TRNSFR FROM ACCOUNT' in d:
        return 'SWEEPS FROM 7459'
    return 'REGULAR DEPOSITS'


def build_excel(transactions, balance_info):
    """Build multi-tab Excel workbook matching the Output 1 format."""
    wb = Workbook()
    money_fmt = '#,##0.00'

    # ── Shared styles ─────────────────────────────────────────────────────────
    header_fill = PatternFill('solid', fgColor=APPLE_BLUE)
    header_font = Font(name='Helvetica', bold=True, size=11, color=WHITE)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_font = Font(name='Helvetica', bold=True, size=14, color=APPLE_DEEP_BLUE)
    section_font = Font(name='Helvetica', bold=True, size=11, color=APPLE_DEEP_BLUE)
    label_font = Font(name='Helvetica', bold=True, size=10, color=APPLE_SEC_GRAY)
    value_font = Font(name='Helvetica', bold=True, size=10, color=APPLE_NEAR_BLACK)
    data_font = Font(name='Helvetica', size=10, color=APPLE_NEAR_BLACK)
    bold_font = Font(name='Helvetica', bold=True, size=10, color=APPLE_NEAR_BLACK)
    red_font = Font(name='Helvetica', bold=True, size=10, color='FF3B30')
    thin_border = Border(bottom=Side(style='thin', color=APPLE_MID_GRAY))
    header_border = Border(bottom=Side(style='medium', color=APPLE_DEEP_BLUE))
    light_gray_fill = PatternFill('solid', fgColor=APPLE_LIGHT_GRAY)
    alt_fill = PatternFill('solid', fgColor='F8F9FA')
    totals_fill = PatternFill('solid', fgColor=APPLE_LIGHT_GRAY)
    totals_border = Border(
        top=Side(style='medium', color=APPLE_DEEP_BLUE),
        bottom=Side(style='medium', color=APPLE_DEEP_BLUE),
    )

    begin_bal = balance_info.get('begin_balance', 0)
    end_bal = balance_info.get('end_balance', 0)
    dep_total_stmt = balance_info.get('deposit_total', 0)
    wth_total_stmt = balance_info.get('withdrawal_total', 0)
    dep_count = balance_info.get('deposit_count', 0)
    wth_count = balance_info.get('withdrawal_count', 0)
    stmt_date = balance_info.get('statement_date', '')
    stmt_thru = balance_info.get('statement_thru', '')
    acct_num = balance_info.get('account_number', '')
    acct_name = balance_info.get('account_name', '')
    acct_type = balance_info.get('account_type', '')
    begin_date = balance_info.get('begin_date', '')
    end_date = balance_info.get('end_date', '')

    # Determine statement month label from begin_date
    if begin_date:
        from datetime import datetime as dt
        try:
            d = dt.strptime(begin_date, '%m/%d/%Y')
            stmt_label = d.strftime('%B %Y')
        except Exception:
            stmt_label = begin_date
    else:
        stmt_label = ''

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1: Statement Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Statement Summary'

    ws['A1'] = 'Atlantic Union Bank'
    ws['A1'].font = title_font
    ws['A2'] = f'Statement – {stmt_label}'
    ws['A2'].font = Font(name='Helvetica', bold=True, size=12, color=APPLE_NEAR_BLACK)

    row = 4
    info_rows = [
        ('Account Number:', acct_num),
        ('Account Name:', acct_name),
        ('Statement Date:', stmt_date),
        ('Statement Thru:', stmt_thru),
        ('Account Type:', acct_type),
    ]
    for lbl, val in info_rows:
        ws.cell(row=row, column=1, value=lbl).font = label_font
        ws.cell(row=row, column=3, value=val).font = value_font
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='BALANCE SUMMARY').font = section_font
    row += 1
    ws.cell(row=row, column=1, value=f'Beginning Balance as of {begin_date}').font = data_font
    ws.cell(row=row, column=3, value=begin_bal).font = value_font
    ws.cell(row=row, column=3).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=1, value=f'+ Deposits and Credits  ({dep_count})').font = data_font
    ws.cell(row=row, column=3, value=dep_total_stmt).font = value_font
    ws.cell(row=row, column=3).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=1, value=f'- Withdrawals and Debits  ({wth_count})').font = data_font
    ws.cell(row=row, column=3, value=wth_total_stmt).font = value_font
    ws.cell(row=row, column=3).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=1, value=f'Ending Balance as of {end_date}').font = bold_font
    ws.cell(row=row, column=3, value=end_bal).font = Font(name='Helvetica', bold=True, size=10, color=APPLE_DEEP_BLUE)
    ws.cell(row=row, column=3).number_format = money_fmt
    row += 1
    for lbl in ['Service Charges for Period', 'Average Collected for Period', 'Minimum Balance for Period']:
        ws.cell(row=row, column=1, value=lbl).font = data_font
        ws.cell(row=row, column=3, value=0).font = data_font
        ws.cell(row=row, column=3).number_format = money_fmt
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='RECONCILIATION CHECK').font = section_font
    row += 1
    recon_start = row
    ws.cell(row=row, column=1, value='Sum of All Deposits (from dep1 sheet)').font = data_font
    ws.cell(row=row, column=3).value = "='Table 1'!C" + str(2 + len(transactions) + 1)
    ws.cell(row=row, column=3).font = value_font
    ws.cell(row=row, column=3).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=1, value='Sum of All Withdrawals (from purch1 sheet)').font = data_font
    ws.cell(row=row, column=3).value = "='Table 1'!D" + str(2 + len(transactions) + 1)
    ws.cell(row=row, column=3).font = value_font
    ws.cell(row=row, column=3).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=1, value='Deposit Difference vs Statement (should be 0)').font = data_font
    ws.cell(row=row, column=3).value = f'=C{recon_start}-{dep_total_stmt}'
    ws.cell(row=row, column=3).font = red_font
    ws.cell(row=row, column=3).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=1, value='Withdrawal Difference vs Statement (should be 0)').font = data_font
    ws.cell(row=row, column=3).value = f'=C{recon_start+1}-{wth_total_stmt}'
    ws.cell(row=row, column=3).font = red_font
    ws.cell(row=row, column=3).number_format = money_fmt

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['C'].width = 18

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2: purch1 (Withdrawals/Debits)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('purch1')

    # Categorize withdrawals
    checks = []
    other_debits = []
    sweeps_to = []
    for t in transactions:
        if t['withdrawal'] is None:
            continue
        cat = categorize_withdrawal(t['description'])
        item = {'date': t['date'], 'description': t['description'], 'amount': t['withdrawal']}
        if cat == 'CHECKS':
            checks.append(item)
        elif cat == 'SWEEPS TO 7459':
            sweeps_to.append(item)
        else:
            other_debits.append(item)

    # Headers
    for col, h in enumerate(['#', 'Date', 'Description', 'Amount', 'Running Total', '', 'Stmt WD Total', 'Difference'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    row = 2
    num = 0

    def write_wd_section(ws, start_row, label, items, num_start):
        r = start_row
        ws.cell(row=r, column=1, value=label).font = section_font
        r += 1
        first_data = r
        for i, item in enumerate(items):
            num = num_start + i + 1
            ws.cell(row=r, column=1, value=num).font = data_font
            ws.cell(row=r, column=2, value=item['date']).font = data_font
            ws.cell(row=r, column=3, value=item['description']).font = data_font
            ws.cell(row=r, column=4, value=item['amount']).font = data_font
            ws.cell(row=r, column=4).number_format = money_fmt
            if r == first_data:
                ws.cell(row=r, column=5).value = f'=D{r}'
            else:
                ws.cell(row=r, column=5).value = f'=D{r}+E{r-1}'
            ws.cell(row=r, column=5).number_format = money_fmt
            ws.cell(row=r, column=5).font = data_font
            if i % 2 == 1:
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = alt_fill
            for c in range(1, 6):
                ws.cell(row=r, column=c).border = thin_border
            r += 1
        last_data = r - 1
        return r, num_start + len(items), first_data, last_data

    # Stmt WD Total in G3
    ws2.cell(row=2, column=7, value='Stmt Total All WD').font = label_font

    # Write CHECKS section
    row, num, chk_first, chk_last = write_wd_section(ws2, row, 'CHECKS', checks, 0)
    ws2.cell(row=3, column=7, value=wth_total_stmt).font = value_font
    ws2.cell(row=3, column=7).number_format = money_fmt

    # Write OTHER DEBITS section
    if other_debits:
        row += 1
        row, num, od_first, od_last = write_wd_section(ws2, row, 'OTHER DEBITS', other_debits, num)

    # Write SWEEPS TO 7459 section
    if sweeps_to:
        row += 1
        row, num, sw_first, sw_last = write_wd_section(ws2, row, 'SWEEPS TO 7459', sweeps_to, num)
        # Subtotal for sweeps
        ws2.cell(row=row, column=3, value='TOTAL SWEEPS TO 7459').font = bold_font
        ws2.cell(row=row, column=4).value = f'=SUM(D{sw_first}:D{sw_last})'
        ws2.cell(row=row, column=4).number_format = money_fmt
        ws2.cell(row=row, column=4).font = bold_font
        ws2.cell(row=row, column=5).value = f'=E{sw_last}'
        ws2.cell(row=row, column=5).number_format = money_fmt
        ws2.cell(row=row, column=5).font = bold_font
        for c in range(1, 6):
            ws2.cell(row=row, column=c).fill = totals_fill
        row += 1

    # Track last individual data row (before any subtotals)
    all_wd_items = checks + other_debits + sweeps_to
    # Count actual data rows written (section labels + items, no subtotals)
    # The last individual item row is: row - 2 if there was a subtotal, otherwise row - 1
    # Simpler: count from row 3 to the row before TOTAL SWEEPS
    last_item_row = row - 2 if sweeps_to else row - 1

    # Grand total
    row += 1
    grand_row = row
    ws2.cell(row=row, column=3, value='GRAND TOTAL ALL WITHDRAWALS').font = Font(
        name='Helvetica', bold=True, size=11, color=APPLE_DEEP_BLUE
    )
    ws2.cell(row=row, column=4).value = f'=SUM(D3:D{last_item_row})'
    ws2.cell(row=row, column=4).number_format = money_fmt
    ws2.cell(row=row, column=4).font = bold_font
    ws2.cell(row=row, column=5).value = f'=D{row}'
    ws2.cell(row=row, column=5).number_format = money_fmt
    ws2.cell(row=row, column=5).font = bold_font
    for c in range(1, 9):
        ws2.cell(row=row, column=c).fill = totals_fill
        ws2.cell(row=row, column=c).border = totals_border

    # Difference
    ws2.cell(row=3, column=8).value = f'=E{grand_row}-G3'
    ws2.cell(row=3, column=8).font = red_font
    ws2.cell(row=3, column=8).number_format = money_fmt

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 16
    ws2.column_dimensions['F'].width = 3
    ws2.column_dimensions['G'].width = 18
    ws2.column_dimensions['H'].width = 16
    ws2.freeze_panes = 'A2'

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3: dep1 (Deposits/Credits)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('dep1')

    regular_deps = []
    sweeps_from = []
    for t in transactions:
        if t['deposit'] is None:
            continue
        cat = categorize_deposit(t['description'])
        item = {'date': t['date'], 'description': t['description'], 'amount': t['deposit']}
        if cat == 'SWEEPS FROM 7459':
            sweeps_from.append(item)
        else:
            regular_deps.append(item)

    for col, h in enumerate(['Date', 'Description', '', 'Amount', 'Running Total', '', 'Stmt Dep Total', 'Difference'], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    row = 2

    def write_dep_section(ws, start_row, label, items):
        r = start_row
        ws.cell(row=r, column=1, value=label).font = section_font
        r += 1
        first_data = r
        for i, item in enumerate(items):
            ws.cell(row=r, column=1, value=item['date']).font = data_font
            ws.cell(row=r, column=2, value=item['description']).font = data_font
            ws.cell(row=r, column=4, value=item['amount']).font = data_font
            ws.cell(row=r, column=4).number_format = money_fmt
            if r == first_data:
                ws.cell(row=r, column=5).value = f'=D{r}'
            else:
                ws.cell(row=r, column=5).value = f'=D{r}+E{r-1}'
            ws.cell(row=r, column=5).number_format = money_fmt
            ws.cell(row=r, column=5).font = data_font
            if i % 2 == 1:
                for c in [1, 2, 4, 5]:
                    ws.cell(row=r, column=c).fill = alt_fill
            for c in [1, 2, 4, 5]:
                ws.cell(row=r, column=c).border = thin_border
            r += 1
        last_data = r - 1
        return r, first_data, last_data

    ws3.cell(row=2, column=7, value='Stmt Total All Deposits').font = label_font

    row, reg_first, reg_last = write_dep_section(ws3, row, 'REGULAR DEPOSITS', regular_deps)
    ws3.cell(row=3, column=7, value=dep_total_stmt).font = value_font
    ws3.cell(row=3, column=7).number_format = money_fmt

    if sweeps_from:
        row += 1
        row, sw_first, sw_last = write_dep_section(ws3, row, 'SWEEPS FROM 7459', sweeps_from)
        ws3.cell(row=row, column=2, value='TOTAL SWEEPS FROM 7459').font = bold_font
        ws3.cell(row=row, column=4).value = f'=SUM(D{sw_first}:D{sw_last})'
        ws3.cell(row=row, column=4).number_format = money_fmt
        ws3.cell(row=row, column=4).font = bold_font
        ws3.cell(row=row, column=5).value = f'=E{sw_last}'
        ws3.cell(row=row, column=5).number_format = money_fmt
        ws3.cell(row=row, column=5).font = bold_font
        for c in [1, 2, 4, 5]:
            ws3.cell(row=row, column=c).fill = totals_fill
        row += 1

    last_dep_item_row = row - 2 if sweeps_from else row - 1

    row += 1
    grand_row = row
    ws3.cell(row=row, column=2, value='GRAND TOTAL ALL DEPOSITS').font = Font(
        name='Helvetica', bold=True, size=11, color=APPLE_DEEP_BLUE
    )
    ws3.cell(row=row, column=4).value = f'=SUM(D3:D{last_dep_item_row})'
    ws3.cell(row=row, column=4).number_format = money_fmt
    ws3.cell(row=row, column=4).font = bold_font
    ws3.cell(row=row, column=5).value = f'=D{row}'
    ws3.cell(row=row, column=5).number_format = money_fmt
    ws3.cell(row=row, column=5).font = bold_font
    for c in range(1, 9):
        ws3.cell(row=row, column=c).fill = totals_fill
        ws3.cell(row=row, column=c).border = totals_border

    ws3.cell(row=3, column=8).value = f'=E{grand_row}-G3'
    ws3.cell(row=3, column=8).font = red_font
    ws3.cell(row=3, column=8).number_format = money_fmt

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 3
    ws3.column_dimensions['D'].width = 16
    ws3.column_dimensions['E'].width = 16
    ws3.column_dimensions['F'].width = 3
    ws3.column_dimensions['G'].width = 18
    ws3.column_dimensions['H'].width = 16
    ws3.freeze_panes = 'A2'

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 4: Table 1 (Full Transaction Detail)
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Table 1')

    for col, h in enumerate(['Date', 'Description', 'Deposits', 'Withdrawals', 'Balance'], 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Beginning balance
    ws4.cell(row=2, column=1, value=begin_date).font = bold_font
    ws4.cell(row=2, column=2, value='BEGINNING BALANCE').font = bold_font
    ws4.cell(row=2, column=5, value=begin_bal).font = bold_font
    ws4.cell(row=2, column=5).number_format = money_fmt
    for c in range(1, 6):
        ws4.cell(row=2, column=c).fill = light_gray_fill
        ws4.cell(row=2, column=c).border = thin_border

    for idx, txn in enumerate(transactions):
        r = 3 + idx
        ws4.cell(row=r, column=1, value=txn['date']).font = data_font
        ws4.cell(row=r, column=2, value=txn['description']).font = data_font
        if txn['deposit'] is not None:
            ws4.cell(row=r, column=3, value=txn['deposit']).font = data_font
            ws4.cell(row=r, column=3).number_format = money_fmt
        if txn['withdrawal'] is not None:
            ws4.cell(row=r, column=4, value=txn['withdrawal']).font = data_font
            ws4.cell(row=r, column=4).number_format = money_fmt
        ws4.cell(row=r, column=5, value=txn['balance']).font = data_font
        ws4.cell(row=r, column=5).number_format = money_fmt
        if idx % 2 == 1:
            for c in range(1, 6):
                ws4.cell(row=r, column=c).fill = alt_fill
        for c in range(1, 6):
            ws4.cell(row=r, column=c).border = thin_border

    last_row = 2 + len(transactions)

    # Totals
    tr = last_row + 1
    ws4.cell(row=tr, column=1, value='TOTALS').font = Font(
        name='Helvetica', bold=True, size=11, color=APPLE_DEEP_BLUE
    )
    ws4.cell(row=tr, column=3).value = f'=SUM(C2:C{last_row})'
    ws4.cell(row=tr, column=3).number_format = money_fmt
    ws4.cell(row=tr, column=3).font = bold_font
    ws4.cell(row=tr, column=4).value = f'=SUM(D2:D{last_row})'
    ws4.cell(row=tr, column=4).number_format = money_fmt
    ws4.cell(row=tr, column=4).font = bold_font
    for c in range(1, 6):
        ws4.cell(row=tr, column=c).fill = totals_fill
        ws4.cell(row=tr, column=c).border = totals_border

    # Reconciliation
    tr += 1
    ws4.cell(row=tr, column=2, value='Statement Deposit Total').font = label_font
    ws4.cell(row=tr, column=3, value=dep_total_stmt).font = value_font
    ws4.cell(row=tr, column=3).number_format = money_fmt
    tr += 1
    ws4.cell(row=tr, column=2, value='Statement Withdrawal Total').font = label_font
    ws4.cell(row=tr, column=4, value=wth_total_stmt).font = value_font
    ws4.cell(row=tr, column=4).number_format = money_fmt
    tr += 1
    ws4.cell(row=tr, column=2, value='Deposit Difference (should be 0)').font = data_font
    ws4.cell(row=tr, column=3).value = f'=C{last_row+1}-{dep_total_stmt}'
    ws4.cell(row=tr, column=3).font = red_font
    ws4.cell(row=tr, column=3).number_format = money_fmt
    tr += 1
    ws4.cell(row=tr, column=2, value='Withdrawal Difference (should be 0)').font = data_font
    ws4.cell(row=tr, column=4).value = f'=D{last_row+1}-{wth_total_stmt}'
    ws4.cell(row=tr, column=4).font = red_font
    ws4.cell(row=tr, column=4).number_format = money_fmt

    ws4.column_dimensions['A'].width = 14
    ws4.column_dimensions['B'].width = 55
    ws4.column_dimensions['C'].width = 18
    ws4.column_dimensions['D'].width = 18
    ws4.column_dimensions['E'].width = 18
    ws4.freeze_panes = 'A2'

    return wb


# ── Streamlit UI ──────────────────────────────────────────────────────────────

uploaded_file = st.file_uploader("Upload Bank Statement PDF", type="pdf")

if uploaded_file:
    if st.button("Convert to Excel", type="primary"):
        with st.spinner("Extracting text from PDF..."):
            pdf_bytes = uploaded_file.read()
            full_text = ""
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        full_text += clean_noise(text) + "\n"

        with st.spinner("Parsing balance summary..."):
            balance_info = extract_balance_summary(full_text)

        with st.spinner("Parsing transactions..."):
            transactions = parse_transactions(full_text)
            begin_bal = balance_info.get('begin_balance', 0)
            transactions = fix_deposit_withdrawal_classification(transactions, begin_bal)

        # ── Display Balance Summary ───────────────────────────────────────
        st.markdown("---")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric(
                "Beginning Balance",
                f"${balance_info.get('begin_balance', 0):,.2f}",
            )
        with col2:
            st.metric(
                "Ending Balance",
                f"${balance_info.get('end_balance', 0):,.2f}",
            )
        with col3:
            st.metric(
                "Deposits",
                f"{balance_info.get('deposit_count', 0)} items",
                f"${balance_info.get('deposit_total', 0):,.2f}",
            )
        with col4:
            st.metric(
                "Withdrawals",
                f"{balance_info.get('withdrawal_count', 0)} items",
                f"${balance_info.get('withdrawal_total', 0):,.2f}",
            )

        # ── Verification ──────────────────────────────────────────────────
        st.markdown("---")
        total_dep = sum(t['deposit'] for t in transactions if t['deposit'])
        total_wth = sum(t['withdrawal'] for t in transactions if t['withdrawal'])
        calc_ending = round(begin_bal + total_dep - total_wth, 2)
        stmt_ending = balance_info.get('end_balance', 0)
        diff = round(calc_ending - stmt_ending, 2)

        vcol1, vcol2, vcol3 = st.columns(3)
        with vcol1:
            st.metric("Calculated Ending Balance", f"${calc_ending:,.2f}")
        with vcol2:
            st.metric("Statement Ending Balance", f"${stmt_ending:,.2f}")
        with vcol3:
            if abs(diff) < 0.02:
                st.metric("Difference", f"${diff:,.2f}", delta="✓ Balanced")
            else:
                st.metric("Difference", f"${diff:,.2f}", delta=f"${diff:,.2f} off", delta_color="inverse")

        st.success(f"Parsed **{len(transactions)}** transactions.")

        # ── Preview ───────────────────────────────────────────────────────
        if transactions:
            preview = []
            for t in transactions[:15]:
                preview.append({
                    'Date': t['date'],
                    'Description': t['description'][:60],
                    'Deposit': f"${t['deposit']:,.2f}" if t['deposit'] else '',
                    'Withdrawal': f"${t['withdrawal']:,.2f}" if t['withdrawal'] else '',
                    'Balance': f"${t['balance']:,.2f}" if t['balance'] else '',
                })
            st.write("**Preview (first 15 transactions):**")
            st.table(preview)

        # ── Build and Download Excel ──────────────────────────────────────
        with st.spinner("Building Excel file..."):
            wb = build_excel(transactions, balance_info)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

        acct = balance_info.get('account_number', 'statement')[-4:]
        stmt_date = balance_info.get('statement_date', '').replace('/', '-')
        filename = f"AUB_Statement_{acct}_{stmt_date}.xlsx"

        st.download_button(
            label="Download Excel File",
            data=output.getvalue(),
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.ml",
            type="primary",
        )
